"""Microsoft Excel (.xlsx / .xlsm) parser — uses openpyxl."""
from __future__ import annotations

import io
from typing import Any

from praxia.io.parsers.base import ParsedFile


class XlsxParser:
    name = "xlsx"

    def parse(
        self,
        data: bytes,
        *,
        filename: str,
        max_rows_per_sheet: int = 200,
        max_cols_per_sheet: int = 50,
        **kwargs: Any,
    ) -> ParsedFile:
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ImportError(
                "Excel (.xlsx) parsing requires `openpyxl`. Install with:\n"
                '  pip install "praxia[office]"'
            ) from e

        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        sections: list[tuple[str, str]] = []
        full_text: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                row_count += 1
                if row_count > max_rows_per_sheet:
                    break
                # Clip excessive columns
                clipped = [
                    "" if v is None else str(v) for v in row[:max_cols_per_sheet]
                ]
                rows.append(clipped)

            if not rows:
                continue

            n_cols = max(len(r) for r in rows)
            # Pad rows to consistent column count
            for r in rows:
                while len(r) < n_cols:
                    r.append("")

            md_lines = []
            md_lines.append("| " + " | ".join(rows[0]) + " |")
            md_lines.append("|" + "|".join(["---"] * n_cols) + "|")
            for row in rows[1:]:
                md_lines.append("| " + " | ".join(row) + " |")
            sheet_md = "\n".join(md_lines)
            sections.append((sheet_name, sheet_md))
            full_text.append(f"\n## Sheet: {sheet_name}\n{sheet_md}")

            if ws.max_row > max_rows_per_sheet:
                full_text.append(
                    f"_(showing first {max_rows_per_sheet} rows of {ws.max_row})_"
                )

        wb.close()

        return ParsedFile(
            filename=filename,
            content="\n".join(full_text),
            metadata={
                "sheet_count": len(wb.sheetnames),
                "sheet_names": wb.sheetnames,
            },
            sections=sections,
        )
